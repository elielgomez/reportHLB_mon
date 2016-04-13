from openpyxl import load_workbook
import warnings
warnings.filterwarnings("ignore")

class Report:
	
	wb = '' 	# Workbook var
	sheet = ''	# Sheet var

	def __init__(this,name):
		this.wb = load_workbook(name) 					# Open file
		this.sheet = this.wb.get_sheet_names()[0]		# Get the first element of the tuple sheet names
	
	def get_edges(this):
	# Gets the Last Column and last Row that have data

		max_row = (this.wb[this.sheet].max_row)
		max_column = (this.wb[this.sheet].max_column)
		return max_row, max_column
	
	def index_fields(this, max_row,max_column):
	# Gets the number of column base on the field name and returns an ASOC ARRAY with the numbers

		fields = {'MUNICIPIO' : 0, 
					'COMUNIDAD' : 0, 
					'SUPERFICIE_HA' : 0, 
					'NO_TRAMPAS_REVISADAS' : 0,
					'NO_TRAMPAS_INSTALADAS' :0 ,
					'NO_DIAPHORINASXTRAMPA' : 0, 
					'PRODUCTOR' : 0,
					'UBICACION' : 0
		}

		for i in range(max_column):
			data = str(this.wb[this.sheet].cell(row = 1, column = i).value)
			for field in fields:
				if field == data.upper():
					fields[field] = i
		return fields

	
	def get_totals(this, index_has, index_mun, index_com, index_trampas, index_instaladas,index_productor, index_ubicacion, index_diaphorina, max_row):
	# It Performs an adittion of all fields in arrays (associative) in one group: Municipio, and subgroup: Localidad (Comunidad)	

		totals = {}

		for i in range(max_row + 1):

			# GET DATA VALUES #

			data_mun = this.wb[this.sheet].cell(row = i, column = index_mun).value 					# Municipio 
			data_com = this.wb[this.sheet].cell(row = i, column = index_com).value 					# Comunidad 
			data_has = this.wb[this.sheet].cell(row = i, column = index_has).value 					# Has
			data_trampas = this.wb[this.sheet].cell(row = i, column = index_trampas).value 			# Trampas
			data_instaladas = this.wb[this.sheet].cell(row = i, column = index_instaladas).value 	# Trampas Instaladas 
			data_productor = this.wb[this.sheet].cell(row = i, column = index_productor).value  	# Productor 
			data_diaphorinas = this.wb[this.sheet].cell(row = i, column = index_diaphorina).value 	# Diaphorina
			data_ubicacion = str(this.wb[this.sheet].cell(row = i, column = index_ubicacion).value)	# Ubicacion (converted to str to get substring)
			data_ubicacion = data_ubicacion[:22]													# Get 22 chars (no trap number)

			if i > 1: # Skip the first Row (Title row)
				if data_mun not in totals: 
					totals[data_mun] = {}

				if data_com not in totals[data_mun]:
					totals[data_mun][data_com] = {} 
					totals[data_mun][data_com]['SUPERFICIE'] = data_has
					totals[data_mun][data_com]['TRAMPAS REVISADAS'] = data_trampas
					totals[data_mun][data_com]['TRAMPAS INSTALADAS'] = data_instaladas
					totals[data_mun][data_com]['DIAPHORINAS'] = data_diaphorinas

					# New arrays to add data and compare
					totals[data_mun][data_com]['UBICACION'] = [] 
					totals[data_mun][data_com]['UBICACION'].append(data_ubicacion)
					totals[data_mun][data_com]['PRODUCTOR'] = []
					totals[data_mun][data_com]['PRODUCTOR'].append(data_productor)
										
				else:
					totals[data_mun][data_com]['SUPERFICIE'] = totals[data_mun][data_com]['SUPERFICIE'] + data_has		
					totals[data_mun][data_com]['TRAMPAS REVISADAS'] = totals[data_mun][data_com]['TRAMPAS REVISADAS'] + data_trampas
					totals[data_mun][data_com]['TRAMPAS INSTALADAS'] = totals[data_mun][data_com]['TRAMPAS INSTALADAS'] + data_instaladas
					totals[data_mun][data_com]['DIAPHORINAS'] = totals[data_mun][data_com]['DIAPHORINAS'] + data_diaphorinas 				
					
					if data_ubicacion not in totals[data_mun][data_com]['UBICACION']:
						totals[data_mun][data_com]['UBICACION'].append(data_ubicacion)
					
					if data_productor not in totals[data_mun][data_com]['PRODUCTOR']:
						totals[data_mun][data_com]['PRODUCTOR'].append(data_productor)
	

		return totals

def main():

	reporte = Report('report.xlsx')						# Instance of Class
	max_row, max_column = reporte.get_edges()			# Get Max Row and Max Column
	index = reporte.index_fields(max_row,max_column)	# Index of each field
	out = 'MUNICIPIO; LOCALIDAD; SUPERFICIE; REVISADAS; INSTALADAS; DIAPHORINAS; PRODUCTORES; PREDIOS \n'	# Out var filled with the titles at first line

	# Call Function to get Data Totals
	data_totals = reporte.get_totals(			
		index['SUPERFICIE_HA'],
		index['MUNICIPIO'],
		index['COMUNIDAD'],
		index['NO_TRAMPAS_REVISADAS'],
		index['NO_TRAMPAS_INSTALADAS'],
		index['PRODUCTOR'],
		index['UBICACION'],
		index['NO_DIAPHORINASXTRAMPA'],
		max_row) 
	for i in sorted(data_totals):
		for j in sorted(data_totals[i]):
			# Set the values in a local var
			mun = i
			loc = j
			sup = data_totals[i][j]['SUPERFICIE']
			rev = data_totals[i][j]['TRAMPAS REVISADAS']
			ins = data_totals[i][j]['TRAMPAS INSTALADAS']
			pro = len(data_totals[i][j]['PRODUCTOR']) # Counts the array len to know the quantity
			ubc = len(data_totals[i][j]['UBICACION']) # Counts the array len to know the quantity
			dph = data_totals[i][j]['DIAPHORINAS']
			out = out + "{0};{1};{2};{3};{4};{5};{6};{7}\n".format(mun,loc,sup,rev,ins,dph,pro,ubc)	# Save the printed data in out

	# Creates a file with the var OUT
	f = open('out.csv','w') 
	f.write(out)

if __name__ == '__main__':
	main()